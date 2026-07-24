from __future__ import annotations

import argparse
import json
from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    "图表编号",
    "优先级",
    "优先级得分",
    "余数编号",
    "数量区间",
    "规格区间",
    "镀种大类",
    "定额单位",
    "样本量",
    "父组系统余数均值",
    "调整判断",
    "P90值",
    "P90预计减少比例",
    "P95值",
    "P95预计减少比例",
    "当前覆盖率",
    "判断来源",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成Excel规则清单和交互HTML")
    parser.add_argument("--input", required=True, type=Path, help="diagnostics.json")
    parser.add_argument("--output-dir", required=True, type=Path)
    return parser.parse_args()


def style_header(cells) -> None:
    for cell in cells:
        cell.fill = PatternFill("solid", fgColor="DDEBF3")
        cell.font = Font(name="Microsoft YaHei", bold=True, color="183B56")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_workbook(payload: dict, output_path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "诊断汇总"
    rule_sheet = workbook.create_sheet("规则调整清单")
    method_sheet = workbook.create_sheet("口径说明")

    summary_sheet.append(["指标", "结果"])
    for key, value in payload["summary"].items():
        summary_sheet.append([key, value])
    style_header(summary_sheet[1])
    summary_sheet.freeze_panes = "A2"
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 22
    for row in summary_sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        label = summary_sheet.cell(row=row[0].row, column=1).value
        if label and ("率" in str(label) or "比例" in str(label)):
            row[0].number_format = "0.00%"
        else:
            row[0].number_format = "#,##0.00"

    rule_sheet.append(HEADERS)
    for record in payload["rules"]:
        rule_sheet.append([record.get(header) for header in HEADERS])
    style_header(rule_sheet[1])
    rule_sheet.freeze_panes = "A2"
    rule_sheet.auto_filter.ref = f"A1:Q{rule_sheet.max_row}"
    widths = [12, 12, 12, 12, 15, 15, 15, 11, 10, 18, 11, 12, 18, 12, 18, 14, 24]
    for index, width in enumerate(widths, start=1):
        rule_sheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(2, rule_sheet.max_row + 1):
        for column in [10, 12, 14]:
            rule_sheet.cell(row, column).number_format = "#,##0.00"
        for column in [13, 15, 16]:
            rule_sheet.cell(row, column).number_format = "0.00%"

    method_rows = [
        ["项目", "口径"],
        ["实际损耗", "领进数量 - 报工数量"],
        ["当前覆盖率", "订单余数 >= 实际损耗的订单占比"],
        ["正向多余余数", "逐单max(余数 - 实际损耗, 0)后求和"],
        ["四维规则", "余数编号×数量区间×规格区间×镀种大类"],
        ["P90/P95", "子组样本量达到门槛时取子组分位数，否则继承三维父组分位数"],
        ["系统均值", "数量组取订单余数均值；百分比组取逐单余数率均值"],
        ["下调建议", "样本量达到门槛，且父组系统余数均值高于P95"],
        ["上调建议", "样本量达到门槛，且父组系统余数均值低于P90"],
        ["维持现状", "父组系统均值位于P90-P95之间，或子组样本量不足"],
        ["P99", "仅影响直方图显示，超过P99的样本不进入绘图箱体，但仍参与规则诊断"],
    ]
    for row in method_rows:
        method_sheet.append(row)
    style_header(method_sheet[1])
    method_sheet.column_dimensions["A"].width = 24
    method_sheet.column_dimensions["B"].width = 80
    for row in method_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    thin = Side(style="thin", color="D9E1E6")
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                font = copy(cell.font)
                font.name = "Microsoft YaHei"
                cell.font = font
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = copy(cell.alignment)
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical or "center",
                    wrap_text=True,
                )
    workbook.save(output_path)


def build_html(payload: dict, output_path: Path) -> None:
    rules_json = json.dumps(payload["rules"], ensure_ascii=False, allow_nan=False)
    html = """<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>余数放量规则实际损耗分布</title>
<style>
:root{--bg:#f5f7f9;--panel:#fff;--text:#1f2933;--muted:#647481;--border:#d9e1e6;--blue:#2374ab;--orange:#d97706;--green:#2f855a}
*{box-sizing:border-box}body{margin:0;background:var(--bg);color:var(--text);font:14px/1.45 "Microsoft YaHei","Segoe UI",sans-serif}
.app{max-width:1450px;margin:auto;padding:20px}.top h1{font-size:24px;margin:0 0 6px}.top p{margin:0 0 16px;color:var(--muted)}
.layout{display:grid;grid-template-columns:320px minmax(0,1fr);gap:16px}.panel{background:#fff;border:1px solid var(--border);border-radius:6px}
.filters{padding:14px}.control{width:100%;height:36px;border:1px solid var(--border);border-radius:5px;padding:0 9px;margin-bottom:10px}
.result-list{max-height:650px;overflow:auto}.result{display:block;width:100%;text-align:left;border:0;border-top:1px solid var(--border);background:#fff;padding:10px;cursor:pointer}
.result.active{background:#e7f1f7;border-left:4px solid var(--blue)}.result span{display:block;color:var(--muted);font-size:12px}
.detail{padding:18px;min-width:0}.detail-head{display:flex;justify-content:space-between}.detail h2{margin:0;font-size:20px}.meta{color:var(--muted);margin:4px 0 14px}
.stats{display:grid;grid-template-columns:repeat(4,minmax(120px,1fr));gap:8px}.stat{border:1px solid var(--border);padding:9px;background:#fafcfd;border-radius:5px}
.stat span{display:block;color:var(--muted);font-size:12px}.legend{display:flex;gap:18px;margin:12px 0;color:var(--muted)}.key{display:flex;align-items:center;gap:6px}.line{width:22px;height:3px}
.chart-wrap{position:relative;height:500px;border:1px solid var(--border);border-radius:5px}.chart{width:100%;height:100%}.tooltip{position:absolute;display:none;background:#17212b;color:#fff;padding:7px 9px;border-radius:4px;pointer-events:none}
.note{margin-top:8px;color:var(--muted)}.btn{height:34px;border:1px solid var(--border);background:#fff;border-radius:5px;padding:0 12px}
@media(max-width:950px){.layout{grid-template-columns:1fr}.result-list{max-height:250px}}@media(max-width:620px){.stats{grid-template-columns:1fr 1fr}}
</style>
</head>
<body>
<main class="app">
<header class="top"><h1>余数放量规则实际损耗分布</h1><p>合成样例 · 图表编号与Excel一致</p></header>
<div class="layout">
<aside class="panel filters">
  <input id="search" class="control" placeholder="输入N0001、余数编号、镀种或区间">
  <div id="count"></div><div id="results" class="result-list"></div>
</aside>
<section class="panel detail">
  <div class="detail-head"><div><h2 id="title"></h2><div id="meta" class="meta"></div></div><div><button id="prev" class="btn">上一条</button><button id="next" class="btn">下一条</button></div></div>
  <div class="stats">
    <div class="stat"><span>优先级 / 得分</span><strong id="priority"></strong></div>
    <div class="stat"><span>样本量</span><strong id="n"></strong></div>
    <div class="stat"><span>父组系统余数均值</span><strong id="system"></strong></div>
    <div class="stat"><span id="quantileLabel">P90 / P95</span><strong id="quantiles"></strong></div>
  </div>
  <div class="legend"><span class="key"><i class="line" style="background:var(--blue)"></i>父组系统均值</span><span class="key"><i class="line" style="background:var(--orange)"></i><span id="legend90">子组P90</span></span><span class="key"><i class="line" style="background:var(--green)"></i><span id="legend95">子组P95</span></span></div>
  <div id="chartWrap" class="chart-wrap"><svg id="chart" class="chart"></svg><div id="tooltip" class="tooltip"></div></div>
  <div id="note" class="note"></div>
</section>
</div>
</main>
<script>
const DATA=__DATA__;
const $=id=>document.getElementById(id),search=$("search"),results=$("results"),svg=$("chart"),tip=$("tooltip"),wrap=$("chartWrap");
let filtered=[...DATA],selected=DATA[0]?.图表编号;
const fmt=(v,d=2)=>Number(v).toLocaleString("zh-CN",{minimumFractionDigits:d,maximumFractionDigits:d});
const suffix=x=>String(x.定额单位).includes("百分")?"%":"百粒";
const inherited=x=>x.分位值来源==="父组";
function apply(){const q=search.value.trim().toLowerCase();filtered=DATA.filter(x=>!q||[x.图表编号,x.余数编号,x.数量区间,x.规格区间,x.镀种大类].join(" ").toLowerCase().includes(q));if(!filtered.some(x=>x.图表编号===selected))selected=filtered[0]?.图表编号;renderList();render()}
function renderList(){$("count").textContent=filtered.length+"条";results.innerHTML=filtered.map(x=>'<button class="result '+(x.图表编号===selected?'active':'')+'" data-id="'+x.图表编号+'"><strong>'+x.图表编号+' · '+x.优先级+' · '+x.调整判断+'</strong><span>'+x.余数编号+' · '+x.镀种大类+' · n='+x.样本量+'</span></button>').join("");results.querySelectorAll("button").forEach(b=>b.onclick=()=>{selected=b.dataset.id;renderList();render()})}
function S(tag,a={}){const e=document.createElementNS("http://www.w3.org/2000/svg",tag);Object.entries(a).forEach(([k,v])=>e.setAttribute(k,v));return e}
function T(x,y,text,a={}){const e=S("text",{x,y,...a});e.textContent=text;return e}
function render(){const x=DATA.find(v=>v.图表编号===selected);if(!x){$("title").textContent="无匹配结果";svg.innerHTML="";return}const parent=inherited(x),su=suffix(x);$("title").textContent=x.图表编号+" · "+x.余数编号+" · "+x.镀种大类;$("meta").textContent="数量区间 "+x.数量区间+" · 规格区间 "+x.规格区间+" · "+x.定额单位+" · "+x.判断来源;$("priority").textContent=x.优先级+" / "+fmt(x.优先级得分,1);$("n").textContent=fmt(x.样本量,0);$("system").textContent=fmt(x.父组系统余数均值)+su;$("quantileLabel").textContent=parent?"父组P90 / 父组P95":"P90 / P95";$("quantiles").textContent=fmt(x.P90值)+su+" / "+fmt(x.P95值)+su;$("legend90").textContent=parent?"父组P90":"子组P90";$("legend95").textContent=parent?"父组P95":"子组P95";$("note").textContent="横轴显示至P99 "+fmt(x.p99)+su+"；绘图排除>P99："+x.excluded+"条，最大值 "+fmt(x.max)+su+"。排除只影响显示。";draw(x);const i=filtered.findIndex(v=>v.图表编号===selected);$("prev").disabled=i<=0;$("next").disabled=i<0||i>=filtered.length-1}
function draw(x){svg.innerHTML="";const W=1000,H=470,m={l:65,r:26,t:78,b:72},pw=W-m.l-m.r,ph=H-m.t-m.b,maxC=Math.max(...x.counts,1),xmin=x.edges[0]??0,xmax=x.p99||1,span=Math.max(xmax-xmin,1e-9),prefix=inherited(x)?"父组":"";svg.setAttribute("viewBox","0 0 "+W+" "+H);for(let i=0;i<=5;i++){const y=m.t+ph-i*ph/5;svg.append(S("line",{x1:m.l,y1:y,x2:W-m.r,y2:y,stroke:"#e5eaee"}),T(m.l-10,y+4,fmt(maxC*i/5,0),{"text-anchor":"end",fill:"#647481","font-size":"12"}))}const bw=pw/x.counts.length;x.counts.forEach((c,i)=>{const h=c/maxC*ph,rect=S("rect",{x:m.l+i*bw+1,y:m.t+ph-h,width:Math.max(1,bw-2),height:h,fill:"#84b7d7"});rect.onmouseenter=e=>{tip.style.display="block";tip.textContent=fmt(x.edges[i])+"–"+fmt(x.edges[i+1])+suffix(x)+"："+c+"条";moveTip(e)};rect.onmousemove=moveTip;rect.onmouseleave=()=>tip.style.display="none";svg.append(rect)});svg.append(S("line",{x1:m.l,y1:m.t+ph,x2:W-m.r,y2:m.t+ph,stroke:"#84929d"}));for(let i=0;i<=6;i++){const value=xmin+span*i/6,px=m.l+pw*i/6;svg.append(T(px,m.t+ph+23,fmt(value),{"text-anchor":"middle",fill:"#647481","font-size":"12"}))}function ref(value,color,label,yoff){const px=m.l+(Math.min(Math.max(value,xmin),xmax)-xmin)/span*pw;svg.append(S("line",{x1:px,y1:m.t,x2:px,y2:m.t+ph,stroke:color,"stroke-width":"3"}),T(Math.min(Math.max(px+6,m.l+6),W-220),m.t-54+yoff,label+" "+fmt(value)+suffix(x),{fill:color,"font-size":"13","font-weight":"600"}))}ref(x.父组系统余数均值,"#2374ab","父组系统均值",0);ref(x.P90值,"#d97706",prefix+"P90",20);ref(x.P95值,"#2f855a",prefix+"P95",40)}
function moveTip(e){const r=wrap.getBoundingClientRect();tip.style.left=Math.min(r.width-220,e.clientX-r.left+12)+"px";tip.style.top=Math.max(5,e.clientY-r.top-38)+"px"}
function nav(d){const i=filtered.findIndex(x=>x.图表编号===selected),n=filtered[i+d];if(n){selected=n.图表编号;renderList();render()}}
search.oninput=apply;$("prev").onclick=()=>nav(-1);$("next").onclick=()=>nav(1);apply();
</script>
</body>
</html>""".replace("__DATA__", rules_json)
    output_path.write_text(html, encoding="utf-8")


def main() -> None:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    payload = json.loads(args.input.read_text(encoding="utf-8"))
    build_workbook(payload, args.output_dir / "rule_diagnostics.xlsx")
    build_html(payload, args.output_dir / "loss_distribution.html")
    print(f"rules={len(payload['rules'])}")
    print(f"xlsx={args.output_dir / 'rule_diagnostics.xlsx'}")
    print(f"html={args.output_dir / 'loss_distribution.html'}")


if __name__ == "__main__":
    main()
